import os
import re
import logging
import openpyxl
from django.conf import settings

logger = logging.getLogger(__name__)

JOBS_DATABASE = []

SUPER_DOMAINS = [
    {"Technology & IT", "Engineering & Manufacturing", "Emerging & Specialized Roles", "Science & Research"},
    {"Business, Finance & Admin", "Sales & Marketing", "Retail & Customer Service", "Legal"},
    {"Hospitality, Food & Tourism", "Beauty & Personal Care"},
    {"Public Safety & Government", "Non-Profit & Community", "Education"},
    {"Skilled Trades & Construction", "Transportation & Logistics", "Agriculture & Environment"}
]

def in_same_super_domain(cat1, cat2):
    if cat1.lower() == cat2.lower():
        return True
    for group in SUPER_DOMAINS:
        group_lower = {g.lower() for g in group}
        if cat1.lower() in group_lower and cat2.lower() in group_lower:
            return True
    return False

def load_jobs_database():
    global JOBS_DATABASE
    if JOBS_DATABASE:
        return
    
    excel_path = os.path.abspath(os.path.join(settings.BASE_DIR, '..', 'Job_Database_AI_Resume_Builder.xlsx'))
    if not os.path.exists(excel_path):
        logger.error(f"Jobs database file not found at: {excel_path}")
        return
        
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_name = 'Jobs Database' if 'Jobs Database' in wb.sheetnames else wb.sheetnames[0]
        sheet = wb[sheet_name]
        
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if row[0]:
                JOBS_DATABASE.append({
                    "id": row[0],
                    "category": row[1] or "General",
                    "title": row[2] or "Specialist",
                    "description": row[3] or "",
                    "skills": [s.strip() for s in row[4].split(",") if s.strip()] if row[4] else [],
                    "education": row[5] or "Varies; Experience-based",
                    "level": row[6] or "Entry"
                })
        logger.info(f"Successfully loaded {len(JOBS_DATABASE)} jobs from excel database.")
    except Exception as e:
        logger.error(f"Error loading jobs database: {e}")

STOP_WORDS = {
    "a", "about", "above", "after", "again", "against", "all", "am", "an", "and", "any", "are", "aren't", "as", "at",
    "be", "because", "been", "before", "being", "below", "between", "both", "but", "by", "can't", "cannot", "could",
    "couldn't", "did", "didn't", "do", "does", "doesn't", "doing", "don't", "down", "during", "each", "few", "for",
    "from", "further", "had", "hadn't", "has", "hasn't", "have", "haven't", "having", "he", "he'd", "he'll", "he's",
    "her", "here", "here's", "hers", "herself", "him", "himself", "his", "how", "how's", "i", "i'd", "i'll", "i'm",
    "i've", "if", "in", "into", "is", "isn't", "it", "it's", "its", "itself", "let's", "me", "more", "most", "mustn't",
    "my", "myself", "no", "nor", "not", "of", "off", "on", "once", "only", "or", "other", "ought", "our", "ours",
    "ourselves", "out", "over", "own", "same", "shan't", "she", "she'd", "she'll", "she's", "should", "shouldn't",
    "so", "some", "such", "than", "that", "that's", "the", "their", "theirs", "them", "themselves", "then", "there",
    "there's", "these", "they", "they'd", "they'll", "they're", "they've", "this", "those", "through", "to", "too",
    "under", "until", "up", "very", "was", "wasn't", "we", "we'd", "we'll", "we're", "we've", "were", "weren't",
    "what", "what's", "when", "when's", "where", "where's", "which", "while", "who", "who's", "whom", "why", "why's",
    "with", "won't", "would", "wouldn't", "you", "you'd", "you'll", "you're", "you've", "your", "yours", "yourself",
    "yourselves", "seeking", "expert", "looking", "experience", "required", "preferred"
}

def get_words(text):
    if not text:
        return set()
    raw_words = re.findall(r'\b[a-z0-9]+\b', text.lower())
    return {w for w in raw_words if w not in STOP_WORDS and len(w) > 1}

def find_best_match(text):
    load_jobs_database()
    words = get_words(text)
    
    fallback_job = {
        "id": "GEN001",
        "category": "General",
        "title": "General Professional",
        "description": "General professional role.",
        "skills": ["Communication", "Problem Solving"],
        "education": "Varies",
        "level": "Mid"
    }

    if not words or not JOBS_DATABASE:
        return fallback_job, 0.0
        
    best_job = None
    best_score = -1.0
    
    TECH_WORDS = {"python", "django", "sql", "api", "node", "react", "cloud", "java", "c++", "git", "code", "html", "css", "developer", "engineer", "software", "development", "programming"}
    has_tech_words = bool(words.intersection(TECH_WORDS))

    for job in JOBS_DATABASE:
        title_words = get_words(job['title'])
        desc_words = get_words(job['description'])
        skills_words = get_words(" ".join(job['skills']))
        
        # Weighted overlap: title is highly critical, skills are important
        title_intersect = len(words.intersection(title_words))
        skills_intersect = len(words.intersection(skills_words))
        desc_intersect = len(words.intersection(desc_words))
        
        score = (title_intersect * 4.0) + (skills_intersect * 2.0) + (desc_intersect * 1.0)
        
        # Tech category tie-breaker boost
        if has_tech_words and job["category"] in {"Technology & IT", "Emerging & Specialized Roles"}:
            score += 0.5
        
        if score > best_score:
            best_score = score
            best_job = job
            
    # If match score is too low (e.g. < 2.0), treat it as General fallback
    if not best_job or best_score < 2.0:
        return fallback_job, best_score if best_job else 0.0
        
    return best_job, best_score
